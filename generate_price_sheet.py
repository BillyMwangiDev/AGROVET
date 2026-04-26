"""
Run: python generate_price_sheet.py
Reads backend/inventory/fixtures/initial_products.json and produces
nicmah_price_update_YYYYMMDD.xlsx ready for price entry + re-import.
"""
import json
from datetime import date
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FIXTURE = Path("backend/inventory/fixtures/initial_products.json")
OUT = Path(f"nicmah_price_update_{date.today().strftime('%Y%m%d')}.xlsx")

data = json.loads(FIXTURE.read_text(encoding="utf-8"))

categories = {r["pk"]: r["fields"]["name"] for r in data if r["model"] == "inventory.category"}
products = sorted(
    [r for r in data if r["model"] == "inventory.product"],
    key=lambda r: (categories.get(r["fields"]["category"], ""), r["fields"]["name"]),
)

# ── Styles ────────────────────────────────────────────────────────────────────
dark = PatternFill("solid", fgColor="0B3A2C")
gold = PatternFill("solid", fgColor="E4B83A")
req_fill = PatternFill("solid", fgColor="FFF3CD")
req_font = Font(name="Calibri", bold=True, color="856404", size=11)
opt_fill = PatternFill("solid", fgColor="E8F5E9")
opt_font = Font(name="Calibri", bold=True, color="2E7D32", size=11)
note_fill = PatternFill("solid", fgColor="D4EDDA")
note_font = Font(name="Calibri", color="155724", size=10)
cat_fill = PatternFill("solid", fgColor="0B3A2C")
cat_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
price_fill = PatternFill("solid", fgColor="FFF9C4")
alt_fill = PatternFill("solid", fgColor="F8FFF9")
wht_fill = PatternFill("solid", fgColor="FFFFFF")
data_font = Font(name="Calibri", size=10)
bs = Side(style="thin", color="DEE2E6")
brd = Border(left=bs, right=bs, top=bs, bottom=bs)
cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEADERS = [
    ("name",          "Product Name *",         "REQUIRED — do not change"),
    ("category",      "Category *",             "REQUIRED — do not change"),
    ("price",         "Selling Price (KES) *",  "FILL IN — number only e.g. 850"),
    ("stock",         "Stock Qty *",            "Current qty on hand"),
    ("unit",          "Unit *",                 "e.g. 50ml, 1kg, bottle"),
    ("description",   "Description",            "Optional"),
    ("supplier",      "Supplier / Brand",       "Optional"),
    ("minstock",      "Min Stock",              "Reorder threshold"),
    ("maxstock",      "Max Stock",              "Max stock target"),
    ("is_ai_product", "Semen Product?",         "YES / NO"),
    ("breed",         "Breed",                  "Semen products only"),
    ("sire_code",     "Sire Code",              "Semen products only"),
]
REQUIRED_KEYS = {"name", "category", "price", "stock", "unit"}
NUM_COLS = len(HEADERS)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

# Row 1 — banner
ws.merge_cells(f"A1:{get_column_letter(NUM_COLS)}1")
c = ws["A1"]
c.value = (
    f"NICMAH AGROVET — Price Update Sheet  •  {date.today().strftime('%d %b %Y')}  "
    "•  Fill in SELLING PRICE for every row, then upload via Admin > Inventory > Import"
)
c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
c.fill = dark
c.alignment = cen
ws.row_dimensions[1].height = 28

# Row 2 — column headers
for ci, (key, label, note) in enumerate(HEADERS, start=1):
    hc = ws.cell(row=2, column=ci, value=label)
    hc.fill = req_fill if key in REQUIRED_KEYS else opt_fill
    hc.font = req_font if key in REQUIRED_KEYS else opt_font
    if key == "price":
        hc.fill = PatternFill("solid", fgColor="F57F17")
        hc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hc.alignment = cen
    hc.border = brd

# Row 3 — notes
for ci, (key, label, note) in enumerate(HEADERS, start=1):
    nc = ws.cell(row=3, column=ci, value=note)
    nc.fill = note_fill
    nc.font = note_font
    nc.alignment = lft
    nc.border = brd

ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 32

# Data rows grouped by category
row = 4
alt = False
prev_cat = None

for p in products:
    f = p["fields"]
    cat_name = categories.get(f["category"], "Unknown")

    # Category header row
    if cat_name != prev_cat:
        ws.merge_cells(f"A{row}:{get_column_letter(NUM_COLS)}{row}")
        cc = ws.cell(row=row, column=1, value=f"▸  {cat_name.upper()}")
        cc.fill = cat_fill
        cc.font = cat_font
        cc.alignment = lft
        cc.border = brd
        ws.row_dimensions[row].height = 20
        row += 1
        prev_cat = cat_name
        alt = False

    fill = alt_fill if alt else wht_fill
    alt = not alt

    price_val = float(f["price"]) if float(f["price"]) != 0.0 else ""

    vals = [
        f["name"],
        cat_name,
        price_val,
        f["stock_level"],
        f["unit"],
        f.get("description", ""),
        f.get("supplier", ""),
        f.get("reorder_point", 5),
        f.get("max_stock", 100),
        "YES" if f.get("is_ai_product") else "NO",
        f.get("breed", ""),
        f.get("sire_code") or "",
    ]

    for ci, val in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = data_font
        cell.border = brd
        cell.alignment = lft
        # Highlight the price column
        if ci == 3:
            cell.fill = price_fill
            cell.font = Font(name="Calibri", size=10, bold=True, color="B71C1C" if not val else "000000")
        else:
            cell.fill = fill

    ws.row_dimensions[row].height = 16
    row += 1

# Column widths
for ci, w in enumerate([40, 24, 20, 12, 14, 35, 22, 12, 12, 16, 20, 18], start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "C4"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
cat_counts = {}
for p in products:
    cn = categories.get(p["fields"]["category"], "Unknown")
    cat_counts[cn] = cat_counts.get(cn, 0) + 1

lines = [
    (f"NICMAH AGROVET — Price Update  {date.today().strftime('%d %b %Y')}", True, "0B3A2C", "FFFFFF"),
    (f"Total products: {len(products)}", False, None, None),
    ("", False, None, None),
    ("HOW TO USE THIS FILE", True, None, None),
    ("  1. Go to the 'Products' sheet.", False, None, None),
    ("  2. Fill in 'Selling Price (KES)' for EVERY row (orange column).", False, None, None),
    ("  3. Optionally update Stock Qty with actual quantities.", False, None, None),
    ("  4. Save as .xlsx", False, None, None),
    ("  5. Admin Dashboard → Inventory → Import Excel → upload this file.", False, None, None),
    ("", False, None, None),
    ("PRODUCTS BY CATEGORY", True, None, None),
]
for cn, count in sorted(cat_counts.items()):
    lines.append((f"  • {cn}  ({count} products)", False, None, None))

for ri, (text, bold, bg, fg) in enumerate(lines, start=1):
    c2 = ws2.cell(row=ri, column=1, value=text)
    c2.font = Font(name="Calibri", bold=bold, color=fg or "111915", size=11 if bold else 10)
    if bg:
        c2.fill = PatternFill("solid", fgColor=bg)
    ws2.row_dimensions[ri].height = 18 if text else 8

ws2.column_dimensions["A"].width = 55

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Products: {len(products)}  |  Categories: {len(cat_counts)}")
print("Rows written:", row - 4 - len(cat_counts), "product rows +", len(cat_counts), "category headers")
