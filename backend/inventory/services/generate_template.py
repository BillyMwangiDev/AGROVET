"""
One-off script: generate nicmah_agrovet_product_import_template.xlsx
Run: python generate_template.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\USER\Downloads\Copy of PRODCTS (1).xlsx"
DST = r"C:\Users\USER\Downloads\nicmah_agrovet_product_import_template.xlsx"

# ── Extract products from source ─────────────────────────────────────────────
wb_src = openpyxl.load_workbook(SRC)
ws_src = wb_src["Sheet1"]

current_company = None
current_desc = None
current_product = None
products = []

for row in ws_src.iter_rows(min_row=2, values_only=True):
    company = row[0]
    desc = row[1]
    product_name = row[3]
    qty = row[5]

    if company:
        current_company = str(company).strip()
    if desc:
        current_desc = str(desc).strip()
    if product_name:
        current_product = str(product_name).strip()

    if current_product and current_desc:
        if qty:
            full_name = f"{current_product} {str(qty).strip()}"
        else:
            full_name = current_product
        products.append({
            "name": full_name.title(),
            "category": current_desc.title(),
            "unit": str(qty).strip() if qty else "unit",
            "supplier": current_company or "",
        })

# Deduplicate
seen = set()
unique = []
for p in products:
    key = (p["name"].lower(), p["category"].lower())
    if key not in seen:
        seen.add(key)
        unique.append(p)

print(f"Unique products extracted: {len(unique)}")

# ── Build template ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products"

# Styles
hdr_fill = PatternFill("solid", fgColor="0B3A2C")
req_fill = PatternFill("solid", fgColor="FFF3CD")
req_font = Font(name="Calibri", bold=True, color="856404", size=11)
opt_fill = PatternFill("solid", fgColor="E8F5E9")
opt_font = Font(name="Calibri", bold=True, color="2E7D32", size=11)
note_fill = PatternFill("solid", fgColor="D4EDDA")
note_font = Font(name="Calibri", color="155724", size=10)
alt_fill = PatternFill("solid", fgColor="F8FFF9")
wht_fill = PatternFill("solid", fgColor="FFFFFF")
data_font = Font(name="Calibri", size=10)
bs = Side(style="thin", color="DEE2E6")
brd = Border(left=bs, right=bs, top=bs, bottom=bs)
cen = Alignment(horizontal="center", vertical="center", wrap_text=True)
lft = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Row 1 — banner
ws.merge_cells("A1:L1")
c = ws["A1"]
c.value = (
    "NICMAH AGROVET — Product Import Template  "
    "•  Fill in PRICE and STOCK for each row, then upload via Admin > Inventory > Import"
)
c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
c.fill = hdr_fill
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Row 2 — column headers
HEADERS = [
    ("name",          "Product Name *",          "REQUIRED — e.g. Alamycin 10 50Ml"),
    ("category",      "Category *",              "REQUIRED — e.g. Injectables, Dewormers"),
    ("price",         "Selling Price (KES) *",   "REQUIRED — number only e.g. 850"),
    ("stock",         "Stock Qty *",             "REQUIRED — current qty e.g. 10"),
    ("unit",          "Unit *",                  "REQUIRED — e.g. 50ml, 1kg, bottle"),
    ("description",   "Description",             "Optional — short product description"),
    ("supplier",      "Supplier / Brand",        "Optional — supplier or brand name"),
    ("minstock",      "Min Stock",               "Optional — reorder threshold (default 5)"),
    ("maxstock",      "Max Stock",               "Optional — max stock target (default 50)"),
    ("is_ai_product", "Semen Product?",          "Optional — YES for bull semen, NO otherwise"),
    ("breed",         "Breed",                   "Optional — breed for semen products"),
    ("sire_code",     "Sire Code",               "Optional — unique sire code"),
]
REQUIRED_KEYS = {"name", "category", "price", "stock", "unit"}

for ci, (key, label, note) in enumerate(HEADERS, start=1):
    hc = ws.cell(row=2, column=ci, value=label)
    hc.fill = req_fill if key in REQUIRED_KEYS else opt_fill
    hc.font = req_font if key in REQUIRED_KEYS else opt_font
    hc.alignment = cen
    hc.border = brd

    nc = ws.cell(row=3, column=ci, value=note)
    nc.fill = note_fill
    nc.font = note_font
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    nc.border = brd

ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 42

# Data rows
for ri, p in enumerate(unique, start=4):
    fill = alt_fill if ri % 2 == 0 else wht_fill
    vals = [
        p["name"],
        p["category"],
        "",          # price — user fills
        0,           # stock
        p["unit"],
        "",          # description
        p["supplier"],
        5,           # minstock
        50,          # maxstock
        "NO",        # is_ai_product
        "",          # breed
        "",          # sire_code
    ]
    for ci, val in enumerate(vals, start=1):
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.font = data_font
        cell.fill = fill
        cell.border = brd
        cell.alignment = lft

# Column widths
for ci, w in enumerate([40, 22, 20, 12, 14, 35, 25, 12, 12, 16, 20, 18], start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "B4"

# ── Instructions sheet ────────────────────────────────────────────────────────
wi = wb.create_sheet("Instructions")
lines = [
    ("NICMAH AGROVET — Import Instructions", True, "0B3A2C", "FFFFFF"),
    ("", False, None, None),
    ("REQUIRED COLUMNS (yellow headers)", True, None, None),
    ("  name          — Full product name", False, None, None),
    ("  category      — Category name (auto-created if new)", False, None, None),
    ("  price         — Selling price in KES (numbers only)", False, None, None),
    ("  stock         — Current stock quantity on hand", False, None, None),
    ("  unit          — Size or unit e.g. 50ml, 1kg, bottle", False, None, None),
    ("", False, None, None),
    ("STEPS", True, None, None),
    ("  1. Open this file and fill in 'Selling Price (KES)' for every row.", False, None, None),
    ("  2. Update 'Stock Qty' with actual quantities you have.", False, None, None),
    ("  3. Optionally add descriptions or adjust min/max stock levels.", False, None, None),
    ("  4. Save the file as .xlsx", False, None, None),
    ("  5. Go to Admin Dashboard > Inventory > Import Excel", False, None, None),
    ("  6. Upload the saved file and click Import.", False, None, None),
    ("", False, None, None),
    ("CATEGORIES IN THIS FILE", True, None, None),
]

categories = sorted(set(p["category"] for p in unique))
for cat in categories:
    count = sum(1 for p in unique if p["category"] == cat)
    lines.append((f"  • {cat} ({count} products)", False, None, None))

for ri, (text, bold, bg, fg) in enumerate(lines, start=1):
    c = wi.cell(row=ri, column=1, value=text)
    color = fg or "111915"
    c.font = Font(name="Calibri", bold=bold, color=color, size=11 if bold else 10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    wi.row_dimensions[ri].height = 18 if text else 8

wi.column_dimensions["A"].width = 70

# Save
wb.save(DST)
print(f"Saved: {DST}")
print(f"Products: {len(unique)}  |  Categories: {len(categories)}")
